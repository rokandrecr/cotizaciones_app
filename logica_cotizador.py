"""
Módulo de lógica de negocio para el sistema de cotizaciones.
Gestiona todas las operaciones relacionadas con el manejo de cotizaciones,
independientemente de la interfaz gráfica.
Incluye funcionalidad para persistir las cotizaciones en formato JSON.
"""

from datetime import datetime
import os
import json

# Intenta importar openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

class GestorCotizaciones:
    """
    Clase que gestiona las cotizaciones, su almacenamiento y manipulación.
    Proporciona métodos para agregar, quitar y filtrar cotizaciones,
    así como para exportarlas a Excel.
    También gestiona la persistencia de las cotizaciones en formato JSON.
    """
    
    # Nombre del archivo para persistencia
    ARCHIVO_COTIZACIONES = "cotizaciones.json"
    
    def __init__(self):
        """
        Inicializa el gestor de cotizaciones.
        Intenta cargar las cotizaciones desde el archivo JSON si existe,
        o inicializa listas vacías si no.
        """
        # Listas para manejar las cotizaciones
        self.cotizaciones_base = []
        self.cotizaciones_disponibles = []
        self.cotizaciones_seleccionadas = []
        
        # Diccionario para almacenar comentarios por cotización
        # La clave es un identificador único (nombre, precio, categoria) y el valor es el comentario
        self.comentarios = {}
        
        # Cargar las cotizaciones desde el archivo JSON si existe
        self.cargar_cotizaciones_desde_json()
    
    def cargar_cotizaciones_desde_json(self):
        """
        Carga las cotizaciones desde el archivo JSON.
        Si el archivo no existe, inicializa las listas vacías.
        
        Returns:
            bool: True si se cargaron las cotizaciones, False si hubo algún error
        """
        try:
            # Verificar si el archivo existe
            if os.path.exists(self.ARCHIVO_COTIZACIONES):
                with open(self.ARCHIVO_COTIZACIONES, 'r', encoding='utf-8') as archivo:
                    # Cargar la lista de diccionarios desde el JSON
                    cotizaciones_json = json.load(archivo)
                    
                    # Convertir de lista de diccionarios a lista de tuplas para uso interno
                    self.cotizaciones_base = []
                    self.comentarios = {}  # Reiniciar comentarios
                    
                    for cotizacion in cotizaciones_json:
                        # Crear tupla de cotización
                        item = (cotizacion["nombre"], cotizacion["precio"], cotizacion["categoria"])
                        self.cotizaciones_base.append(item)
                        
                        # Guardar comentario si existe
                        if "comentario" in cotizacion and cotizacion["comentario"]:
                            self.comentarios[item] = cotizacion["comentario"]
                    
                    # Inicializar las listas de disponibles y seleccionadas
                    self.cotizaciones_disponibles = list(self.cotizaciones_base)
                    self.cotizaciones_seleccionadas = []
                    
                    print(f"Cotizaciones cargadas desde {self.ARCHIVO_COTIZACIONES}: {len(self.cotizaciones_base)} items")
                    return True
            else:
                # Si el archivo no existe, inicializar con listas vacías
                self.cotizaciones_base = []
                self.cotizaciones_disponibles = []
                self.cotizaciones_seleccionadas = []
                self.comentarios = {}
                print(f"Archivo {self.ARCHIVO_COTIZACIONES} no encontrado. Se inicia con listas vacías.")
                return True
                
        except json.JSONDecodeError as e:
            print(f"Error al decodificar el archivo JSON: {e}")
            # Si hay un error en el formato JSON, inicializar con listas vacías
            self.cotizaciones_base = []
            self.cotizaciones_disponibles = []
            self.cotizaciones_seleccionadas = []
            self.comentarios = {}
            return False
            
        except Exception as e:
            print(f"Error al cargar las cotizaciones: {e}")
            # En caso de cualquier otro error, inicializar con listas vacías
            self.cotizaciones_base = []
            self.cotizaciones_disponibles = []
            self.cotizaciones_seleccionadas = []
            self.comentarios = {}
            return False
    
    def guardar_cotizaciones_en_json(self):
        """
        Guarda las cotizaciones en un archivo JSON.
        
        Returns:
            bool: True si se guardaron las cotizaciones, False si hubo algún error
        """
        try:
            # Convertir de lista de tuplas a lista de diccionarios para el JSON
            cotizaciones_json = []
            
            for item in self.cotizaciones_base:
                nombre, precio, categoria = item
                
                # Crear diccionario base con los campos obligatorios
                cotizacion_dict = {
                    "nombre": nombre, 
                    "precio": precio, 
                    "categoria": categoria
                }
                
                # Añadir el comentario si existe
                if item in self.comentarios:
                    cotizacion_dict["comentario"] = self.comentarios[item]
                else:
                    cotizacion_dict["comentario"] = ""  # Guardar campo vacío por defecto
                
                cotizaciones_json.append(cotizacion_dict)
            
            # Guardar en el archivo JSON
            with open(self.ARCHIVO_COTIZACIONES, 'w', encoding='utf-8') as archivo:
                json.dump(cotizaciones_json, archivo, ensure_ascii=False, indent=2)
                
            print(f"Cotizaciones guardadas en {self.ARCHIVO_COTIZACIONES}: {len(self.cotizaciones_base)} items")
            return True
            
        except Exception as e:
            print(f"Error al guardar las cotizaciones: {e}")
            return False
            
    def cargar_cotizaciones_iniciales(self, cotizaciones_iniciales=None):
        """
        Carga las cotizaciones iniciales al sistema.
        Si ya existen cotizaciones cargadas desde el JSON, no hace nada.
        
        Args:
            cotizaciones_iniciales: Lista opcional de tuplas (nombre, precio, categoria)
        """
        # Si ya hay cotizaciones cargadas, no hacer nada
        if self.cotizaciones_base:
            return
            
        # Si se proporcionan cotizaciones iniciales, utilizarlas
        if cotizaciones_iniciales:
            self.cotizaciones_base = list(cotizaciones_iniciales)
            self.cotizaciones_disponibles = list(self.cotizaciones_base)
            self.cotizaciones_seleccionadas = []
            
            # Guardar las cotizaciones iniciales en el JSON
            self.guardar_cotizaciones_en_json()
        # Si ya hay cotizaciones cargadas, no hacer nada
        if self.cotizaciones_base:
            return
            
        # Si se proporcionan cotizaciones iniciales, utilizarlas
        if cotizaciones_iniciales:
            self.cotizaciones_base = list(cotizaciones_iniciales)
            self.cotizaciones_disponibles = list(self.cotizaciones_base)
            self.cotizaciones_seleccionadas = []
            
            # Guardar las cotizaciones iniciales en el JSON
            self.guardar_cotizaciones_en_json()
    
    def obtener_categorias_unicas(self):
        """
        Obtiene la lista de categorías únicas en las cotizaciones base.
        
        Returns:
            list: Lista ordenada de categorías únicas
        """
        return sorted(set(categoria for _, _, categoria in self.cotizaciones_base))
    
    def filtrar_disponibles_por_categoria(self, categoria):
        """
        Filtra las cotizaciones disponibles por categoría.
        
        Args:
            categoria: Categoría por la que filtrar, o "Todas" para mostrar todas
            
        Returns:
            list: Lista filtrada de cotizaciones disponibles
        """
        # Si la categoría es "Todas", mostrar todas las cotizaciones
        if categoria == "Todas":
            self.cotizaciones_disponibles = [(nombre, precio, cat) 
                                           for nombre, precio, cat in self.cotizaciones_base
                                           if (nombre, precio, cat) not in self.cotizaciones_seleccionadas]
        else:
            # Filtrar por la categoría seleccionada
            self.cotizaciones_disponibles = [(nombre, precio, cat) 
                                           for nombre, precio, cat in self.cotizaciones_base
                                           if cat == categoria and (nombre, precio, cat) not in self.cotizaciones_seleccionadas]
        
        return self.cotizaciones_disponibles
    
    def agregar_a_seleccionadas(self, item):
        """
        Agrega un ítem a las cotizaciones seleccionadas y lo quita de las disponibles.
        
        Args:
            item: Tupla (nombre, precio, categoria) a agregar
            
        Returns:
            bool: True si se agregó correctamente, False si ya existía
        """
        if item in self.cotizaciones_seleccionadas:
            return False
        
        self.cotizaciones_seleccionadas.append(item)
        if item in self.cotizaciones_disponibles:
            self.cotizaciones_disponibles.remove(item)
        
        # No es necesario guardar en JSON aquí, ya que solo se mueve un ítem entre listas
        # y la base de cotizaciones no cambia
        return True
    
    def obtener_comentario(self, item):
        """
        Obtiene el comentario asociado a una cotización.
        
        Args:
            item: Tupla (nombre, precio, categoria) de la cotización
            
        Returns:
            str: El comentario asociado o cadena vacía si no existe
        """
        return self.comentarios.get(item, "")
    
    def establecer_comentario(self, item, comentario):
        """
        Establece o actualiza el comentario para una cotización.
        
        Args:
            item: Tupla (nombre, precio, categoria) de la cotización
            comentario: Texto del comentario a establecer
            
        Returns:
            bool: True si se estableció correctamente
        """
        self.comentarios[item] = comentario
        
        # Guardar los cambios en el archivo JSON
        return self.guardar_cotizaciones_en_json()
    
    def quitar_de_seleccionadas(self, item):
        """
        Quita un ítem de las cotizaciones seleccionadas y lo devuelve a las disponibles.
        
        Args:
            item: Tupla (nombre, precio, categoria) a quitar
            
        Returns:
            bool: True si se quitó correctamente, False si no existía
        """
        if item not in self.cotizaciones_seleccionadas:
            return False
        
        self.cotizaciones_seleccionadas.remove(item)
        if item in self.cotizaciones_base and item not in self.cotizaciones_disponibles:
            self.cotizaciones_disponibles.append(item)
        
        # No es necesario guardar en JSON aquí, ya que solo se mueve un ítem entre listas
        # y la base de cotizaciones no cambia
        return True
    
    def calcular_total(self):
        """
        Calcula el total de las cotizaciones seleccionadas.
        
        Returns:
            float: Suma de los precios de las cotizaciones seleccionadas
        """
        return sum(precio for _, precio, _ in self.cotizaciones_seleccionadas)
    
    def nueva_cotizacion(self):
        """
        Reinicia la cotización actual, moviendo todos los ítems a disponibles.
        """
        # Restaurar todas las cotizaciones a disponibles
        self.cotizaciones_disponibles = list(self.cotizaciones_base)
        self.cotizaciones_seleccionadas = []
        
        # No es necesario guardar en JSON aquí, ya que solo se reinician las listas
        # y la base de cotizaciones no cambia
    
    def agregar_nueva_cotizacion_base(self, nombre, precio, categoria):
        """
        Agrega una nueva cotización a la lista base y la hace disponible.
        También guarda automáticamente todas las cotizaciones en el archivo JSON.
        
        Args:
            nombre: Nombre o descripción de la cotización
            precio: Precio de la cotización
            categoria: Categoría de la cotización
            
        Returns:
            tuple: La nueva cotización agregada como (nombre, precio, categoria)
        """
        nueva_cotizacion = (nombre, precio, categoria)
        
        # Agregar a las listas
        self.cotizaciones_base.append(nueva_cotizacion)
        self.cotizaciones_disponibles.append(nueva_cotizacion)
        
        # Guardar los cambios en el archivo JSON
        self.guardar_cotizaciones_en_json()
        
        return nueva_cotizacion
    
    def obtener_comentario(self, item):
        """
        Obtiene el comentario asociado a una cotización.
        
        Args:
            item: Tupla (nombre, precio, categoria) de la cotización
            
        Returns:
            str: El comentario asociado o cadena vacía si no existe
        """
        return self.comentarios.get(item, "")
    
    def establecer_comentario(self, item, comentario):
        """
        Establece o actualiza el comentario para una cotización.
        
        Args:
            item: Tupla (nombre, precio, categoria) de la cotización
            comentario: Texto del comentario a establecer
            
        Returns:
            bool: True si se estableció correctamente
        """
        self.comentarios[item] = comentario
        
        # Guardar los cambios en el archivo JSON
        return self.guardar_cotizaciones_en_json()
    
    def eliminar_cotizacion_base(self, item):
        """
        Elimina una cotización de la lista base y de cualquier otra lista donde esté.
        También guarda automáticamente el estado en el archivo JSON.
        
        Args:
            item: Tupla (nombre, precio, categoria) a eliminar
            
        Returns:
            bool: True si se eliminó correctamente, False si no existía
        """
        if item not in self.cotizaciones_base:
            return False
        
        # Eliminar de todas las listas
        self.cotizaciones_base.remove(item)
        
        if item in self.cotizaciones_disponibles:
            self.cotizaciones_disponibles.remove(item)
            
        if item in self.cotizaciones_seleccionadas:
            self.cotizaciones_seleccionadas.remove(item)
        
        # Eliminar comentario si existe
        if item in self.comentarios:
            del self.comentarios[item]
        
        # Guardar los cambios en el archivo JSON
        self.guardar_cotizaciones_en_json()
        
        return True
    
    def exportar_a_excel(self):
        """
        Exporta la cotización actual a un archivo Excel.
        
        Returns:
            dict: Diccionario con información sobre el resultado de la exportación
                  {
                      'exito': True/False,
                      'archivo': 'ruta al archivo' (si exito=True),
                      'mensaje': 'mensaje de error' (si exito=False)
                  }
        """
        if not OPENPYXL_DISPONIBLE:
            return {
                "exito": False,
                "mensaje": "El módulo 'openpyxl' no está instalado."
            }
        
        if not self.cotizaciones_seleccionadas:
            return {
                "exito": False,
                "mensaje": "No hay items en la cotización actual."
            }
        
        try:
            # Crear un nuevo libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cotización"
            
            # Definir estilos
            estilo_encabezado = Font(bold=True, size=12)
            estilo_total = Font(bold=True, size=12)
            borde = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            alineacion_centro = Alignment(horizontal='center')
            alineacion_derecha = Alignment(horizontal='right')
            
            # Añadir encabezados
            encabezados = ["Descripción", "Precio (CRC)", "Categoría"]
            for col, encabezado in enumerate(encabezados, start=1):
                celda = ws.cell(row=1, column=col, value=encabezado)
                celda.font = estilo_encabezado
                celda.alignment = alineacion_centro
                celda.border = borde
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 40  # Descripción
            ws.column_dimensions['B'].width = 15  # Precio
            ws.column_dimensions['C'].width = 25  # Categoría
            
            # Añadir los datos de las cotizaciones
            for idx, (nombre, precio, categoria) in enumerate(self.cotizaciones_seleccionadas, start=2):
                # Descripción
                celda = ws.cell(row=idx, column=1, value=nombre)
                celda.border = borde
                
                # Precio - formateo especial para números
                celda = ws.cell(row=idx, column=2, value=precio)
                celda.number_format = '$#,##0.00'
                celda.alignment = alineacion_derecha
                celda.border = borde
                
                # Categoría
                celda = ws.cell(row=idx, column=3, value=categoria)
                celda.border = borde
            
            # Calcular el total
            fila_total = len(self.cotizaciones_seleccionadas) + 2
            total = self.calcular_total()
            
            # Celda "Total:"
            celda = ws.cell(row=fila_total, column=1, value="TOTAL:")
            celda.font = estilo_total
            celda.alignment = alineacion_derecha
            celda.border = borde
            
            # Celda con el valor del total
            celda = ws.cell(row=fila_total, column=2, value=total)
            celda.font = estilo_total
            celda.number_format = '$#,##0.00'
            celda.alignment = alineacion_derecha
            celda.border = borde
            
            # Celda vacía en la columna de categoría (para completar la fila del total)
            celda = ws.cell(row=fila_total, column=3, value="")
            celda.border = borde
            
            # Generar nombre de archivo con fecha y hora
            fecha_hora = datetime.now().strftime("%Y%m%d_%H%M")
            nombre_archivo = f"cotizacion_{fecha_hora}.xlsx"
            
            # Guardar el archivo en el directorio actual
            ruta_completa = os.path.join(os.getcwd(), nombre_archivo)
            wb.save(ruta_completa)
            
            return {
                "exito": True,
                "archivo": ruta_completa
            }
            
        except Exception as e:
            # Capturar cualquier error que pueda ocurrir
            return {
                "exito": False,
                "mensaje": str(e)
            }